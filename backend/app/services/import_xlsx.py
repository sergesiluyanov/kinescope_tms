"""Импорт тест-кейсов из xlsx-экспорта Test IT.

Структура файла (наблюдаемая в экспорте Test IT):
- Заголовок: ID, Расположение, Наименование, Автоматизирован, Предусловия,
  Шаги, Постусловия, Ожидаемый результат, Тестовые данные, Комментарии,
  Итерации, Приоритет, Статус, Дата создания, Автор, Длительность, Тег.
- Каждый кейс — это строка с непустым ID, за которой следуют 1..N строк с
  пустыми ID, в которых указаны конкретные шаги/предусловия/тег.

Парсер устойчив к смещению колонок: он находит индексы по русским
заголовкам, поэтому строгий порядок не требуется.
"""

from __future__ import annotations

import io
import re
from dataclasses import dataclass, field
from typing import Any

from openpyxl import load_workbook

from app.models.test_case import TestCasePriority

# ----- Маппинги -----

# Шаги внутри одной ячейки часто пронумерованы — ловим оба варианта:
# `1. ...` и `1) ...`. Требуем именно начало строки, чтобы не сматчить
# версии вида «1.0.0» или «п.1» в середине текста.
INLINE_STEP_RE = re.compile(r"^[ \t]*(\d+)[\.\)][ \t]+", re.MULTILINE)


PRIORITY_MAP: dict[str, TestCasePriority] = {
    "самый низкий": TestCasePriority.low,
    "низкий": TestCasePriority.low,
    "средний": TestCasePriority.medium,
    "высокий": TestCasePriority.high,
    "самый высокий": TestCasePriority.critical,
    "critical": TestCasePriority.critical,
    "high": TestCasePriority.high,
    "medium": TestCasePriority.medium,
    "low": TestCasePriority.low,
    # Короткая нотация «P0..P3», часто встречается в шаблонах руками собранных:
    # P0 — must-have / blocker, P3 — nice-to-have.
    "p0": TestCasePriority.critical,
    "p1": TestCasePriority.high,
    "p2": TestCasePriority.medium,
    "p3": TestCasePriority.low,
}

REQUIRED_HEADERS: dict[str, list[str]] = {
    "id": ["ID"],
    # Полный путь раздела одной строкой («Раздел -> Подраздел -> ...»).
    "location": ["Расположение", "Раздел", "Path", "Folder Path"],
    # Альтернатива: раздел и подраздел в разных колонках. Если есть «section»
    # и «subsection», иерархия собирается из них в порядке section -> subsection.
    "section": ["Section", "Раздел верхнего уровня", "Folder", "Suite"],
    "subsection": ["Subsection", "Подраздел", "Папка", "Subfolder"],
    "title": ["Наименование", "Название", "Заголовок", "Title"],
    "preconditions": ["Предусловия", "Preconditions", "Pre-conditions"],
    "steps": ["Шаги", "Steps"],
    "expected": ["Ожидаемый результат", "Expected", "Expected Result"],
    "priority": ["Приоритет", "Priority"],
    "tag": ["Тег", "Теги", "Tag", "Tags"],
    # Произвольные заметки — кладём в preconditions с префиксом, чтобы
    # не терять информацию, которая часто оказывается важной для воспроизведения.
    "notes": ["Notes", "Заметки", "Комментарии", "Comments"],
}

SECTION_SEPARATOR_RE = re.compile(r"\s*->\s*|\s*/\s*")


@dataclass
class ParsedStep:
    action: str
    expected: str

    def to_dict(self) -> dict[str, str]:
        return {"action": self.action, "expected": self.expected}


@dataclass
class ParsedCase:
    external_id: str | None
    section_path: list[str]
    title: str
    preconditions: str | None
    steps: list[ParsedStep]
    priority: TestCasePriority
    tags: list[str]
    # «Общий ожидаемый результат» — отдельная строка без action.
    # Используем для заголовка, а если основной заголовок уже есть — приклеим
    # к предусловиям, чтобы информация не потерялась.
    general_expected: str | None = None
    raw_status: str | None = None
    # Сколько ячеек оригинального файла пошло в этот кейс — полезно для UI.
    source_rows: int = 0


@dataclass
class SectionStat:
    path: list[str]
    new: bool


@dataclass
class ImportPlan:
    cases: list[ParsedCase]
    section_paths: list[list[str]] = field(default_factory=list)
    issues: list[str] = field(default_factory=list)

    @property
    def total_cases(self) -> int:
        return len(self.cases)


# ----- Утилиты -----


def _norm(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _truthy(v: Any) -> bool:
    return _norm(v) != ""


def _resolve_priority(raw: str) -> TestCasePriority:
    return PRIORITY_MAP.get(raw.strip().lower(), TestCasePriority.medium)


def _split_section_path(raw: str) -> list[str]:
    parts = [p.strip() for p in SECTION_SEPARATOR_RE.split(raw) if p.strip()]
    return parts


def split_numbered_block(text: str) -> list[str]:
    """Если в ячейке лежит пронумерованный список («1. ...\\n2. ...»),
    разбить его на отдельные пункты. Промежуточные строки без номера
    (вложенные буллеты, дополнения) приклеиваются к предыдущему пункту.

    Если нумерации нет (или только один номер) — возвращаем исходный текст
    одним блоком; если текст пустой — пустой список.
    """

    if not text or not text.strip():
        return []
    matches = list(INLINE_STEP_RE.finditer(text))
    # Меньше двух номеров — это не список, оставляем как есть.
    if len(matches) < 2:
        return [text.strip()]

    parts: list[str] = []
    # Если перед первым номером есть «вступительный» текст — сохраним его
    # как отдельный нулевой пункт, чтобы ничего не потерять.
    if matches[0].start() > 0:
        preamble = text[: matches[0].start()].strip()
        if preamble:
            parts.append(preamble)

    for i, m in enumerate(matches):
        body_start = m.end()
        body_end = matches[i + 1].start() if i + 1 < len(matches) else len(text)
        body = text[body_start:body_end].strip()
        if body:
            parts.append(body)
    return parts


def _build_title(
    case_external_id: str | None,
    title_cell: str,
    general_expected: str | None,
) -> str:
    """Заголовок может быть пустым в Test IT-экспорте — берём первый
    осмысленный фрагмент. Если осмысленного нет — генерируем уникальный
    идентификатор, чтобы не плодить одинаковые заголовки."""

    if title_cell:
        return title_cell.strip()[:300]

    if general_expected:
        return general_expected.strip().splitlines()[0][:300]

    if case_external_id:
        return f"Кейс #{case_external_id}"

    return "Без названия"


# ----- Основной парсер -----


def _resolve_columns(header: tuple[Any, ...]) -> dict[str, int]:
    """Маппим логические поля на индексы колонок по заголовкам."""

    by_name = {_norm(h).lower(): idx for idx, h in enumerate(header) if h is not None}
    result: dict[str, int] = {}
    for key, candidates in REQUIRED_HEADERS.items():
        for c in candidates:
            idx = by_name.get(c.lower())
            if idx is not None:
                result[key] = idx
                break
    return result


def parse_xlsx_bytes(
    content: bytes,
    *,
    drop_root_section: bool = False,
    split_inline_steps: bool = True,
) -> ImportPlan:
    """Превратить байты xlsx в список кейсов.

    `drop_root_section`: убрать первый уровень из «Расположения» — удобно
    для Test IT-экспортов, у которых верхний уровень совпадает с именем
    проекта (например, всегда «Kinescope»).

    `split_inline_steps`: если в одной ячейке «Шаги» лежит пронумерованный
    список вида «1. ... \\n2. ...», разрезать его на отдельные шаги.
    «Ожидаемый результат» в Test IT почти всегда один на всю
    последовательность — в этом случае он навешивается на последний шаг.
    Если expected тоже пронумерован и совпадает по числу пунктов —
    распределим попарно.
    """

    wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return ImportPlan(cases=[], issues=["Файл пустой"])

    header = rows[0]
    cols = _resolve_columns(header)
    issues: list[str] = []
    # Единственная по-настоящему обязательная колонка — это «Шаги»: без
    # действий кейс бесполезен. Иерархия раздела опциональна (если нет —
    # кейсы упадут в раздел «Импорт»); ID опционален тоже — без него
    # каждая строка = отдельный кейс (форма «один кейс на строку»).
    if "steps" not in cols:
        issues.append(
            "В файле не нашлась колонка с шагами. "
            "Поддерживаются заголовки: «Шаги» или «Steps»."
        )
        return ImportPlan(cases=[], issues=issues)
    has_id_column = "id" in cols
    has_location_column = "location" in cols
    has_section_columns = "section" in cols or "subsection" in cols

    def cell(row: tuple[Any, ...], key: str) -> str:
        idx = cols.get(key)
        if idx is None or idx >= len(row):
            return ""
        return _norm(row[idx])

    cases: list[ParsedCase] = []
    current: ParsedCase | None = None
    current_pre_lines: list[str] = []
    current_tags_set: set[str] = set()
    current_step_buffers: list[ParsedStep] = []
    current_general_expected: list[str] = []

    def flush_current() -> None:
        if current is None:
            return
        pre_parts = list(current_pre_lines)
        general_expected = (
            "\n".join(current_general_expected).strip()
            if current_general_expected
            else None
        )
        # Заголовок в Test IT часто пустой — пробуем достать осмысленный.
        new_title = _build_title(
            current.external_id, current.title, general_expected
        )
        # Если заголовок оказался «техническим» (`Кейс #...`) и есть общий
        # ожидаемый — оставим его в preconditions, чтобы информация не пропала.
        used_general_as_title = (
            general_expected is not None
            and new_title.strip() == general_expected.strip().splitlines()[0][:300]
        )
        if general_expected and not used_general_as_title:
            pre_parts.append(f"Ожидаемый результат: {general_expected}")

        current.preconditions = "\n".join(pre_parts).strip() or None
        current.steps = current_step_buffers
        current.tags = sorted(current_tags_set)
        current.general_expected = general_expected
        current.title = new_title
        cases.append(current)

    def resolve_section_path(row: tuple[Any, ...]) -> list[str]:
        """Иерархия раздела из «Расположения» (приоритет) либо из
        пары «Section / Subsection». Если нет ни того, ни другого —
        пустой путь (кейс попадёт в дефолтный раздел «Импорт»)."""
        if has_location_column:
            location_raw = cell(row, "location")
            if location_raw:
                return _split_section_path(location_raw)
        if has_section_columns:
            parts: list[str] = []
            sec_val = cell(row, "section")
            if sec_val:
                parts.append(sec_val.strip())
            sub_val = cell(row, "subsection")
            if sub_val:
                parts.append(sub_val.strip())
            return parts
        return []

    def is_new_case_row(row: tuple[Any, ...]) -> bool:
        """Считаем, что начался новый кейс, если:
        - В файле есть колонка ID и в ней непустое значение, ИЛИ
        - Колонки ID нет вовсе — тогда каждая строка с непустыми
          «Шагами» или «Названием» считается отдельным кейсом."""
        if has_id_column:
            return bool(cell(row, "id"))
        return bool(cell(row, "steps") or cell(row, "title"))

    for row in rows[1:]:
        if is_new_case_row(row):
            # Новый кейс — закрыть предыдущий.
            flush_current()
            ext_id = cell(row, "id") if has_id_column else None
            section_path = resolve_section_path(row)
            if drop_root_section and len(section_path) > 1:
                section_path = section_path[1:]
            current = ParsedCase(
                external_id=ext_id or None,
                section_path=section_path,
                title=cell(row, "title"),
                preconditions=None,
                steps=[],
                priority=_resolve_priority(cell(row, "priority")),
                tags=[],
                raw_status=cell(row, "status") if "status" in cols else None,
                source_rows=1,
            )
            current_pre_lines = []
            current_tags_set = set()
            current_step_buffers = []
            current_general_expected = []

            # Предусловия, шаги, теги и заметки могут лежать в этой же
            # строке (форма «один кейс на строку»). В Test IT-форме
            # «шаги расписаны построчно» они появятся в следующих строках
            # и попадут в ту же логику ниже — без дублирования: после
            # continue мы уйдём сюда же на следующей итерации только если
            # is_new_case_row снова true.
            pre = cell(row, "preconditions")
            if pre:
                current_pre_lines.append(pre)

            notes = cell(row, "notes") if "notes" in cols else ""
            if notes:
                current_pre_lines.append(f"Заметки: {notes}")

            action = cell(row, "steps")
            expected = cell(row, "expected")
            if action:
                sub_actions = (
                    split_numbered_block(action)
                    if split_inline_steps
                    else [action]
                )
                if len(sub_actions) <= 1:
                    current_step_buffers.append(
                        ParsedStep(action=action, expected=expected)
                    )
                else:
                    sub_expecteds = (
                        split_numbered_block(expected)
                        if split_inline_steps and expected
                        else []
                    )
                    if len(sub_expecteds) == len(sub_actions):
                        for a, e in zip(sub_actions, sub_expecteds, strict=False):
                            current_step_buffers.append(
                                ParsedStep(action=a, expected=e)
                            )
                    else:
                        last_idx = len(sub_actions) - 1
                        for idx, a in enumerate(sub_actions):
                            current_step_buffers.append(
                                ParsedStep(
                                    action=a,
                                    expected=expected if idx == last_idx else "",
                                )
                            )
            elif expected:
                current_general_expected.append(expected)

            tag_val = cell(row, "tag")
            if tag_val:
                for t in re.split(r"[,\n;]+", tag_val):
                    t = t.strip().lstrip("#").strip()
                    if t:
                        current_tags_set.add(t)
            continue

        if current is None:
            # Пропускаем «осиротевшие» строки до первого ID.
            continue

        current.source_rows += 1

        pre = cell(row, "preconditions")
        if pre:
            current_pre_lines.append(pre)

        notes = cell(row, "notes") if "notes" in cols else ""
        if notes:
            current_pre_lines.append(f"Заметки: {notes}")

        action = cell(row, "steps")
        expected = cell(row, "expected")
        if action:
            sub_actions = (
                split_numbered_block(action) if split_inline_steps else [action]
            )
            if len(sub_actions) <= 1:
                current_step_buffers.append(
                    ParsedStep(action=action, expected=expected)
                )
            else:
                sub_expecteds = (
                    split_numbered_block(expected) if split_inline_steps and expected else []
                )
                if len(sub_expecteds) == len(sub_actions):
                    # «Шаг N» ↔ «Ожидаемое N» — попарно.
                    for a, e in zip(sub_actions, sub_expecteds, strict=False):
                        current_step_buffers.append(ParsedStep(action=a, expected=e))
                else:
                    # Один общий ожидаемый результат на всю последовательность
                    # — это «итог сценария», вешаем его на последний шаг.
                    last_idx = len(sub_actions) - 1
                    for idx, a in enumerate(sub_actions):
                        current_step_buffers.append(
                            ParsedStep(
                                action=a,
                                expected=expected if idx == last_idx else "",
                            )
                        )
        elif expected:
            # «Общий» ожидаемый результат — нет действия, только expected.
            current_general_expected.append(expected)

        tag_val = cell(row, "tag")
        if tag_val:
            for t in re.split(r"[,\n;]+", tag_val):
                t = t.strip().lstrip("#").strip()
                if t:
                    current_tags_set.add(t)

    flush_current()

    # Собираем уникальные пути для UI.
    section_paths_set: set[tuple[str, ...]] = set()
    for c in cases:
        for i in range(1, len(c.section_path) + 1):
            section_paths_set.add(tuple(c.section_path[:i]))
    section_paths = sorted(list(p) for p in section_paths_set)

    return ImportPlan(cases=cases, section_paths=section_paths, issues=issues)


# ----- Запись в БД -----


async def commit_import(
    db,  # AsyncSession (избегаем import чтобы файл оставался лёгким)
    *,
    project_id: int,
    plan: ImportPlan,
    created_by_id: int | None,
) -> dict[str, int]:
    """Сохраняет план в БД. Возвращает счётчики."""
    from sqlalchemy import select

    from app.models.section import Section
    from app.models.test_case import TestCase, TestCaseStatus

    # Подгружаем все секции проекта одной выборкой.
    existing_q = await db.execute(
        select(Section).where(Section.project_id == project_id)
    )
    sections = list(existing_q.scalars().all())

    # Индекс: parent_id (None для корня) + lower(name) → section.
    section_by_key: dict[tuple[int | None, str], Section] = {
        (s.parent_id, s.name.strip().lower()): s for s in sections
    }
    # Чтобы корректно проставлять position в новых разделах.
    next_position: dict[int | None, int] = {}
    for s in sections:
        prev = next_position.get(s.parent_id, -1)
        if s.position > prev:
            next_position[s.parent_id] = s.position

    sections_created = 0

    async def ensure_path(path: list[str]) -> Section | None:
        nonlocal sections_created
        parent_id: int | None = None
        last: Section | None = None
        for name in path:
            key = (parent_id, name.strip().lower())
            sec = section_by_key.get(key)
            if sec is None:
                pos = next_position.get(parent_id, -1) + 1
                next_position[parent_id] = pos
                sec = Section(
                    project_id=project_id,
                    parent_id=parent_id,
                    name=name,
                    description=None,
                    position=pos,
                )
                db.add(sec)
                await db.flush()
                section_by_key[key] = sec
                sections_created += 1
            last = sec
            parent_id = sec.id
        return last

    # Для кейсов без указанного «расположения» создадим раздел «Импорт».
    fallback: Section | None = None

    cases_created = 0
    for case in plan.cases:
        if case.section_path:
            target = await ensure_path(case.section_path)
        else:
            if fallback is None:
                fallback = await ensure_path(["Импорт"])
            target = fallback

        if target is None:
            continue

        tc = TestCase(
            section_id=target.id,
            title=case.title,
            preconditions=case.preconditions,
            steps=[s.to_dict() for s in case.steps],
            priority=case.priority.value,
            status=TestCaseStatus.active.value,
            tags=list(case.tags),
            created_by_id=created_by_id,
        )
        db.add(tc)
        cases_created += 1

    await db.commit()
    return {
        "cases_created": cases_created,
        "sections_created": sections_created,
    }
